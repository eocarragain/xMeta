import scenario
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.parse import urlparse

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "links"
header = ['page_url', 'url', 'final_url', 'host', 'ext', 'link_type', 'text', "final_status", "first_status", "wayback", "wayback_status", "wayback_first_status", "pages_encountered"]
out_ws.append(header)

unique_dict = {}

wb = load_workbook("scen_toc.xlsx")
ws =  wb.get_sheet_by_name("TOC")
for row in ws.iter_rows():
    url = row[1].value 
    if url == "article_url":
        continue

    art = scenario.parseScenario(url)
    links = art.get_html_links()
    for link_dict in links:
        link = links[link_dict]
        url = link["url"]
        host = urlparse(url).netloc
        url_end = url.split("/")[-1]
        if "." in url_end:
            ext = url_end.split(".")[-1]
        else:
            ext = ""
        if url in unique_dict:
            unique_dict[url]["pages_found"].append(link["page_url"]) 
        else:
            unique_dict[url] = {
                "first_page_encountered" : link["page_url"],
                "url" : url,
                "final_url": link["final_url"],
                "host" : host,
                "ext": ext,
                "text": link["text"],
                "link_type": link["link_type"],
                "url_status": link["url_status"],
                "first_status":  link["first_status"],
                "wayback_url": link["wayback"],
                "wayback_status":  link["wayback_status"],
                "wayback_first_status" : link["wayback_first_status"],
                "pages_found": [link["page_url"]]
            }

print(unique_dict)

for link_key in unique_dict:
    link = unique_dict[link_key]
    out_ws.append([link["first_page_encountered"], link['url'], link['final_url'], link['host'], link['ext'], link["link_type"], link["text"], link["url_status"], link["first_status"], link["wayback_url"], link["wayback_status"], link["wayback_first_status"], '; '.join(link["pages_found"])])

out_wb.save("scenario_links.xlsx")
