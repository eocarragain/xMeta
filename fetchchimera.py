import scenario
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import re
from unidecode import unidecode
import datetime
import fetchutils

article_urls = [
    "http://research.ucc.ie/journals/chimera/2013/00/kukuliac/03/en",
    "http://research.ucc.ie/journals/chimera/2013/00/glynn/02/en",
    "http://research.ucc.ie/journals/chimera/2013/00/mccroy/04/en",
    "http://research.ucc.ie/journals/chimera/2013/00/osullivan/05/en",
    "http://research.ucc.ie/journals/chimera/2013/00/sheridanquantz/06/en",
    "http://research.ucc.ie/journals/chimera/2013/00/williams/07/en",
    "http://research.ucc.ie/journals/chimera/2013/00/scriven/08/en",
    "http://research.ucc.ie/journals/chimera/2013/00/kandrot/09/en",
    "http://research.ucc.ie/journals/chimera/2013/00/lennon/11/en"
]

#issue_urls = ["http://research.ucc.ie/boolean/2010/00"]


if __name__ == '__main__':
    utils = fetchutils.fetchUtils("contribs_chimera.xlsx")
    base_doi = '10.33178/chimera'
    default_affil = "University College Cork, Ireland."

    wb = Workbook()

    # grab the active worksheet
    journal_ws = wb.active
    journal_ws.title = "Journal"
    journal_header = ['doi', 'journal_title', 'short_title', 'journal_issn', 'journal_url', 'reference_distribution_opts', 'publisher', 'license_url']
    journal_values = [base_doi, 'Chimera','Chimera', '', 'https://research.ucc.ie/chimera/home', 'any', 'University College Cork', '']
    journal_ws.append(journal_header)
    journal_ws.append(journal_values)

    volume_ws = wb.create_sheet("Volume")
    volume_header = ['doi', 'volume_number', 'volume_url']
    volume_values = ['', '26', '']
    volume_ws.append(volume_header)
    volume_ws.append(volume_values)

    issue_ws = wb.create_sheet("Issue")
    issue_header = ['doi', 'issue_title', 'editors', 'publication_date', 'issue_number', 'issue_url', 'collection', 'rights_statement','languages'] 
    issue_doi = "{0}.26".format(base_doi)
    issue_values = [
        issue_doi,
        '',
        utils.get_contibs(['Richard Scriven'], default_affil),
        '2013-09-11',
        '2012/2013',
        'http://research.ucc.ie/journals/chimera/home',
        '',
        '',
        'en'
    ]
    issue_ws.append(issue_header)
    issue_ws.append(issue_values)
    issue_ws['D1'].number_format = 'yyyy-mm-dd'

    articles_ws = wb.create_sheet("Articles")
    art_header = ['doi','language','title','subtitle','authors','editors','publication_date','url','abstract','abstract_translated_to_en','first_page','last_page','keywords','keywords_de','type','peer_reviewed','Recommended citation for item', "mint_doi"]
    articles_ws.append(art_header)
    
    citations_ws = wb.create_sheet("Citations")
    citations_header = ["Article DOI ","Complete reference","DOI for reference"]
    citations_ws.append(citations_header)

    for article_url in article_urls:
        url_parts = article_url.split("/")
        art_id = str(int(url_parts[-2]))
        art_doi = "{0}.{1}".format(issue_doi, art_id)
        language = 'en'
        art = scenario.parseChimera(article_url)
        title = art.title
        status_code = art.get_status_code()

        #Get affiliation. Only one per article
        affil = art.get_affiliation(default_affil)
        authors = art.get_authors('Breffní Lennon')
        keywords = art.get_keywords()
        author_keys = utils.get_contibs(authors, affil)
        if len(author_keys.strip()) == 0:
            raise Exception("No authors found when fetching {}".format(art_doi))
        art_values = [
            art_doi,
            language,
            title,
            '',
            author_keys, #todo
            utils.get_contibs(['Richard Scriven'], default_affil), #todo
            '2013-09-11',#utils.get_full_date(art.get_meta_tag("citation_publication_date")),
            article_url,
            art.get_abstract(),
            '',
            art.get_start_page(),
            art.get_end_page(),# art.get_meta_tag("citation_lastpage"),
            keywords,# keywords
            '',# keywords_de
            'Article',# type
            'Non peer-reviewed',# peer_reviewed
            '', # recommended citation
            'True' # skip doi
        ]
        
        citations = art.get_citations()
        for citation in citations:
            citations_ws.append([art_doi, citation, ''])

        articles_ws.append(art_values)

    contributors_ws = wb.create_sheet("Contributors")
    contributors_header = ['id', 'given_name', 'surname', 'orcid', 'primary_affiliation', 'secondary_affiliation', 'email_for_ucc_authors']
    contributors_ws.append(contributors_header)
    for r in dataframe_to_rows(utils.contrib_df, index=False, header=False):
        contributors_ws.append(r)

    wb.save("chimera_2013.xlsx")
    utils.save_wb()
        
    