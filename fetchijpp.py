import scenario
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import re
from unidecode import unidecode
import datetime
import fetchutils

issue_urls = [
    #"http://research.ucc.ie/ijpp/2009/01",
    #"http://research.ucc.ie/ijpp/2010/01",
    "http://research.ucc.ie/ijpp/2011/01",
    "http://research.ucc.ie/ijpp/2011/02",
    "http://research.ucc.ie/ijpp/2012/01"
]

#issue_urls = ["http://research.ucc.ie/ijpp/2010/01"]


if __name__ == '__main__':
    utils = fetchutils.fetchUtils("contribs_ijpp.xlsx")
    default_affil = "University College Cork, Ireland."
    base_doi = '10.33178/ijpp'
    for url in issue_urls:
        url_parts = url.split('/')
        year = url_parts[-2]
        issue_no = url_parts[-1]
        issue_no_as_int = str(int(issue_no))

        issue = scenario.parseIjppIssue(url)
        print(issue.issue_url)

        vol = issue.get_volume()
        vol_as_int = issue.get_volume_as_int(vol)

        wb = Workbook()

        # grab the active worksheet
        journal_ws = wb.active
        journal_ws.title = "Journal"
        journal_header = ['doi', 'journal_title', 'short_title', 'journal_issn', 'journal_url', 'reference_distribution_opts', 'publisher', 'license_url']
        journal_values = [base_doi, 'Irish Journal of Public Policy','Irish Journal of Public Policy', '', 'https://research.ucc.ie/ijpp', 'any', 'University College Cork', '']
        journal_ws.append(journal_header)
        journal_ws.append(journal_values)
 
        volume_ws = wb.create_sheet("Volume")
        volume_header = ['doi', 'volume_number', 'volume_url']
        volume_values = ['', vol, '']
        volume_ws.append(volume_header)
        volume_ws.append(volume_values)

        issue_ws = wb.create_sheet("Issue")
        issue_header = ['doi', 'issue_title', 'editors', 'publication_date', 'issue_number', 'issue_url', 'collection', 'rights_statement','languages'] 
        issue_doi = "{0}.{1}.{2}".format(base_doi, vol_as_int, issue_no_as_int)
        issue_title = "Volume {} Issue {} ({})".format(str(vol), issue_no_as_int, year)
        issue_values = [
            issue_doi,
            issue_title,
            utils.get_contibs(issue.get_editors(), default_affil),
            utils.get_full_date(issue.get_year(), issue_no),
            issue_no_as_int,
            url,
            '',
            'en'
        ]
        issue_ws.append(issue_header)
        issue_ws.append(issue_values)
        issue_ws['D1'].number_format = 'yyyy-mm-dd'

        articles_ws = wb.create_sheet("Articles")
        art_header = ['doi','language','title','subtitle','authors','editors','publication_date','url','abstract','abstract_translated_to_en','first_page','last_page','keywords','keywords_de','type','peer_reviewed','Recommended citation for item', "mint_doi"]
        articles_ws.append(art_header)
        
        citations_ws = wb.create_sheet("Citations")
        citations_header = ["Article DOI ","Complete reference","DOI for reference"]
        citations_ws.append(citations_header)

        article_urls = issue.get_unique_article_urls()
        #print("######## arts from tocs: {}".format(issue.get_articles_from_toc()['http://research.ucc.ie/ijpp/2009/01/otuama/00/en']))
        for article_url in article_urls:
            if "http" not in article_url:
                article_url = "{0}/{1}".format(url.rsplit("/", 1)[0], article_url)
            url_parts = article_url.split("/")
            art_id = str(int(url_parts[-2]))
            art_doi = "{0}.{1}".format(issue_doi, art_id)
            language = url_parts[-1]
            art = scenario.parseIjpp(article_url)
            title = art.title
            status_code = art.get_status_code()

            #Get affiliation. Only one per article
            affil = default_affil
            # todo handle affiliaions!!!


            #toc_section = art.get_section()
            #section_ref = issue.get_section_ref(toc_section, title)
            #non_ojs_meta = issue.get_section_meta_for_non_ojs(section_ref)
            
            # return a dict of authors and affils
            authors = art.get_authors()
            auth_affils = art.get_authors_affils()
            print(auth_affils)
            author_keys = ""
            for auth_affil in auth_affils:
                auth = auth_affils[auth_affil]['name']
                affil = auth_affils[auth_affil]['affiliation']
                print(auth)
                print(affil)
                auth_key = utils.get_contibs([auth], affil)
                print(auth_key)
                if author_keys == "":
                    author_keys = auth_key
                else:
                    author_keys = "{0}||{1}".format(author_keys, auth_key)

            #author_keys = utils.get_contibs(authors, affil)
            if len(author_keys.strip()) == 0:
                print("@@@@@ {}".format(authors))
                raise Exception("No authors found when fetching {}".format(art_doi))
            art_values = [
                art_doi,
                language,
                title,
                '',
                author_keys, #todo
                utils.get_contibs(issue.get_editors(), default_affil), #todo
                '',#utils.get_full_date(art.get_meta_tag("citation_publication_date")),
                article_url,
                art.get_abstract(),
                '',
                art.get_start_page(),
                art.get_end_page(),# art.get_meta_tag("citation_lastpage"),
                '',# keywords
                '',# keywords_de
                'Article',# type
                'Non peer-reviewed',# peer_reviewed
                '', # recommended citation
                'True' # skip doi
            ]
            
            citations = art.get_citations()
            for citation in citations:
                citations_ws.append([art_doi, citation, ''])

            articles_ws.append(art_values)

        contributors_ws = wb.create_sheet("Contributors")
        contributors_header = ['id', 'given_name', 'surname', 'orcid', 'primary_affiliation', 'secondary_affiliation', 'email_for_ucc_authors']
        contributors_ws.append(contributors_header)
        for r in dataframe_to_rows(utils.contrib_df, index=False, header=False):
            contributors_ws.append(r)

        wb.save("ijpp_{}_{}.xlsx".format(year, issue_no))
        utils.save_wb()
        
    